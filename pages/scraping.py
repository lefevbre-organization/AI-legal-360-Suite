import streamlit as st
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import io
import logging
from time import sleep

# Configurar logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Agregar Google Fonts y aplicar estilos personalizados
st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;600&display=swap');
    body { font-family: 'Poppins', sans-serif; }
    h1, h2, h3, h4, h5, h6 { color: #001978; }
    .log-text { white-space: pre-wrap; word-wrap: break-word; }
    </style>
    """,
    unsafe_allow_html=True
)

# Configuración de la aplicación
st.markdown("#### Buscador de Palabras Clave en Sitios Web")

# Entrada de palabras clave
keywords_input = st.text_area("Palabras clave (separadas por comas)", 
    "denuncia, denuncias, canal de denuncias, canal, canal ético, compliance, "
    "Channel, ethics, complaint, canaldenuncias, canaletico, etico, ético, "
    "código de conducta, code of conduct, whistleblower channel, Reporting channel, "
    "Whistleblowing channel, canal de ética, ética, Complaints Channel, "
    "Sistema Interno de Información, Canal del informante, Canal de información, "
    "Canal de comunicación interno, General conditions of sale")

keywords = [kw.strip().lower() for kw in keywords_input.split(",")]

# Subir archivo Excel
uploaded_file = st.file_uploader("Sube tu archivo Excel", type=["xlsx"])

# Configuración de columna
selected_column = st.text_input("Nombre de la columna de links", "WEBSITE")

# Headers para simular un navegador
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
}

def make_request(url, retries=3, delay=5):
    session = requests.Session()
    session.headers.update(headers)
    for _ in range(retries):
        try:
            response = session.get(url, timeout=10, verify=False)
            response.raise_for_status()
            return response
        except requests.exceptions.RequestException as e:
            logging.error(f"Error al acceder a {url}: {e}")
            sleep(delay)
            delay *= 2
    return None

# Botón de ejecución
log_display = st.empty()

st.write("---")

if st.button("Ejecutar búsqueda") and uploaded_file and selected_column:
    try:
        workbook = load_workbook(uploaded_file)
        sheet = workbook.active

        website_column_index = None
        for cell in sheet[1]:
            if cell.value == selected_column:
                website_column_index = cell.column
                break

        if not website_column_index:
            st.error(f"Columna '{selected_column}' no encontrada.")
            st.stop()

        result_col_index = sheet.max_column + 1
        sheet.cell(row=1, column=result_col_index, value="Resultado")

        log_entries = []

        for i, row in enumerate(range(2, sheet.max_row + 1), start=1):
            url = sheet.cell(row=row, column=website_column_index).value
            if url:
                if not url.startswith("http"):
                    url = f"https://{url}"
                try:
                    response = make_request(url)
                    if response:
                        soup = BeautifulSoup(response.content, "html.parser")
                        text = soup.get_text().lower()

                        found = False
                        for keyword in keywords:
                            if keyword in text:
                                link = soup.find('a', string=lambda text: text and keyword in text.lower())
                                link_href = link['href'] if link else f"Palabra clave encontrada: '{keyword}' - Link no encontrado"
                                sheet.cell(row=row, column=result_col_index, value=link_href)
                                log_entries.append(f"{i}: ✔️ Palabra clave '{keyword}' encontrada en {url}")
                                found = True
                                break

                        if not found:
                            sheet.cell(row=row, column=result_col_index, value="Palabras clave no encontradas")
                            log_entries.append(f"{i}: ⚠️ No se encontraron palabras clave en {url}")
                    else:
                        sheet.cell(row=row, column=result_col_index, value="Error al acceder después de reintentos")
                        log_entries.append(f"{i}: ❌ Error al acceder a {url}")
                except Exception as e:
                    sheet.cell(row=row, column=result_col_index, value=f"Error inesperado: {e}")
                    log_entries.append(f"{i}: ❌ Error inesperado en {url}: {e}")
            else:
                sheet.cell(row=row, column=result_col_index, value="URL vacía")
                log_entries.append(f"{i}: ⚠️ URL vacía en la fila {row}")

            log_display.markdown('<div class="log-text">' + '\n'.join(log_entries[-10:]) + '</div>', unsafe_allow_html=True)

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        st.success("Archivo procesado con éxito.")
        st.download_button(
            label="Descargar archivo procesado",
            data=output,
            file_name="output_with_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"Ocurrió un error: {e}")
